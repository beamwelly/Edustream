import openpyxl

wb = openpyxl.load_workbook('/home/akhila/Downloads/Learning_Platform_App/Needs_Discovery_latest (1) (1).xlsx', data_only=False)

def inspect_sheet(sheet_name):
    print("=" * 60)
    print(f"SHEET: {sheet_name}")
    print("=" * 60)
    sheet = wb[sheet_name]
    for r in range(1, sheet.max_row + 1):
        row_vals = []
        has_content = False
        for c in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=r, column=c)
            val = cell.value
            if val is not None:
                has_content = True
            # Format nicely
            coord = cell.coordinate
            row_vals.append((coord, val))
        if has_content:
            line = []
            for coord, val in row_vals:
                if val is not None:
                    line.append(f"{coord}: {val}")
            print(" | ".join(line))

for name in wb.sheetnames:
    inspect_sheet(name)
