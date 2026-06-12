from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Form
from fastapi.responses import JSONResponse
from pydantic import BaseModel
from typing import List, Optional
from datetime import datetime
import uuid
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, update, delete

from app.database.session import get_db
from app.models.user import User
from app.models.tool import ToolRegistry
from app.utils.security import decode_access_token
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from app.services.notification_service import create_notification
from app.utils.supabase_storage import upload_file_to_supabase, download_file_from_supabase, BUCKET_NAME, SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY

# Calculator service imports
from app.services.wow.retirement_calculator import calculate_retirement_predictor
from app.services.wow.cost_delay_calculator import calculate_cost_of_delay
from app.services.wow.sip_home_loan_calculator import calculate_sip_home_loan_impact
from app.services.wow.freedom_date_calculator import calculate_freedom_date
from app.services.wow.goal_dashboard_calculator import calculate_goal_dashboard

router = APIRouter(prefix="/wow", tags=["WOW Financial Freedom Toolkit"])
security = HTTPBearer()

# --- Authentication Dependencies ---

async def get_current_active_user(
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: AsyncSession = Depends(get_db)
) -> User:
    token = credentials.credentials
    payload = decode_access_token(token)
    if not payload:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Session expired or invalid token."
        )
    user_id = payload.get("user_id")
    if not user_id:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Token is missing required user identity claims."
        )
    stmt = select(User).where(User.id == user_id)
    result = await db.execute(stmt)
    user = result.scalar_one_or_none()
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="User associated with this session no longer exists."
        )
    if not user.is_active:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="User account is deactivated."
        )
    if user.role == "employee":
        from app.models.employee_permission import EmployeePermission
        perm_stmt = select(EmployeePermission).where(EmployeePermission.user_id == user.id)
        perm_res = await db.execute(perm_stmt)
        perm = perm_res.scalar_one_or_none()
        if not perm or "wow_toolkit" not in (perm.allowed_tools or []):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Access Denied. You do not have permission to access the WOW Toolkit."
            )
    return user

async def get_current_admin(
    current_user: User = Depends(get_current_active_user)
) -> User:
    if current_user.role != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Forbidden. This action requires Admin privileges."
        )
    return current_user


# --- Tool Registry Pydantic Schemas ---

class ToolResponse(BaseModel):
    id: int
    name: str
    description: str
    type: str  # "interactive", "downloadable"
    file_path: Optional[str] = None
    original_filename: Optional[str] = None
    storage_filename: Optional[str] = None
    icon_name: str
    is_active: bool
    created_at: datetime
    updated_at: datetime

    class Config:
        from_attributes = True

class ToolCreate(BaseModel):
    name: str
    description: str
    type: str
    file_path: Optional[str] = None
    original_filename: Optional[str] = None
    storage_filename: Optional[str] = None
    icon_name: str
    is_active: bool = True

class ToolUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    type: Optional[str] = None
    file_path: Optional[str] = None
    original_filename: Optional[str] = None
    storage_filename: Optional[str] = None
    icon_name: Optional[str] = None
    is_active: Optional[bool] = None


# --- Tool Registry API Endpoints ---

@router.get("/tools", response_model=List[ToolResponse])
async def list_tools(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    # Fetch all tools
    stmt = select(ToolRegistry)
    res = await db.execute(stmt)
    tools = res.scalars().all()

    # Self-healing default seed if registry is completely empty
    if not tools:
        seed_tools = [
            ToolRegistry(
                name="WOW Financial Freedom Toolkit",
                description="Comprehensive financial independence workbook containing retirement age predictor, cost of delay calculator, and family financial vault.",
                type="interactive",
                icon_name="Coins",
                is_active=True
            ),
            ToolRegistry(
                name="Retirement Planner (Excel)",
                description="Offline Excel-based calculator to estimate corpus requirements for a comfortable retirement.",
                type="downloadable",
                icon_name="PiggyBank",
                file_path=f"{SUPABASE_URL}/storage/v1/object/public/{BUCKET_NAME}/tools/Retirement_Planner.xlsx",
                original_filename="Retirement_Planner.xlsx",
                storage_filename="Retirement_Planner.xlsx",
                is_active=True
            )
        ]
        db.add_all(seed_tools)
        await db.commit()
        
        stmt = select(ToolRegistry)
        res = await db.execute(stmt)
        tools = res.scalars().all()

    # Filter visible tools for standard users
    if current_user.role == "admin":
        return tools
    return [t for t in tools if t.is_active]

@router.post("/tools")
async def create_tool(
    payload: ToolCreate,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(get_current_admin)
):
    print(f"[DATABASE_TOOL_CREATE] Attempting to create tool: '{payload.name}'")
    try:
        # Pre-validate unique constraint (Option B)
        stmt = select(ToolRegistry).where(ToolRegistry.name == payload.name)
        res = await db.execute(stmt)
        if res.scalar_one_or_none():
            print(f"[DATABASE_TOOL_CREATE] Name validation failed: tool '{payload.name}' already exists.")
            return JSONResponse(
                status_code=status.HTTP_409_CONFLICT,
                content={"success": False, "message": "A tool with this name already exists."}
            )

        tool = ToolRegistry(**payload.model_dump())
        db.add(tool)
        await db.commit()
        await db.refresh(tool)
        print(f"[DATABASE_TOOL_CREATE] Tool '{payload.name}' successfully registered with ID: {tool.id}")

        # Notify users
        await create_notification(
            db=db,
            title="New Tool Available",
            message=f"{tool.name} is now available.",
            type="tool_available",
            reference_id=str(tool.id),
            target_group="users"
        )

        return tool
    except Exception as e:
        await db.rollback()
        print(f"[DATABASE_TOOL_CREATE] Integrity or Database exception occurred: {str(e)}")
        if "unique constraint" in str(e).lower() or "ix_tools_registry_name" in str(e).lower():
            return JSONResponse(
                status_code=status.HTTP_409_CONFLICT,
                content={"success": False, "message": "A tool with this name already exists."}
            )
        return JSONResponse(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            content={"success": False, "message": f"Database error: {str(e)}"}
        )

@router.put("/tools/{tool_id}")
async def update_tool(
    tool_id: int,
    payload: ToolUpdate,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(get_current_admin)
):
    print(f"[DATABASE_TOOL_CREATE] Attempting to update tool ID: {tool_id}")
    try:
        stmt = select(ToolRegistry).where(ToolRegistry.id == tool_id)
        res = await db.execute(stmt)
        tool = res.scalar_one_or_none()
        if not tool:
            return JSONResponse(
                status_code=status.HTTP_404_NOT_FOUND,
                content={"success": False, "message": "Tool not found"}
            )
            
        payload_dict = payload.model_dump(exclude_unset=True)
        # Pre-validate unique constraint on rename
        if "name" in payload_dict and payload_dict["name"] != tool.name:
            stmt_clash = select(ToolRegistry).where(ToolRegistry.name == payload_dict["name"])
            res_clash = await db.execute(stmt_clash)
            if res_clash.scalar_one_or_none():
                print(f"[DATABASE_TOOL_CREATE] Rename validation failed: tool '{payload_dict['name']}' already exists.")
                return JSONResponse(
                    status_code=status.HTTP_409_CONFLICT,
                    content={"success": False, "message": "A tool with this name already exists."}
                )
                
        for k, v in payload_dict.items():
            setattr(tool, k, v)
            
        await db.commit()
        await db.refresh(tool)
        print(f"[DATABASE_TOOL_CREATE] Tool ID: {tool_id} updated successfully.")
        return tool
    except Exception as e:
        await db.rollback()
        print(f"[DATABASE_TOOL_CREATE] Exception updating tool ID: {tool_id}: {str(e)}")
        if "unique constraint" in str(e).lower() or "ix_tools_registry_name" in str(e).lower():
            return JSONResponse(
                status_code=status.HTTP_409_CONFLICT,
                content={"success": False, "message": "A tool with this name already exists."}
            )
        return JSONResponse(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            content={"success": False, "message": f"Database error: {str(e)}"}
        )

@router.delete("/tools/{tool_id}")
async def delete_tool(
    tool_id: int,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(get_current_admin)
):
    print(f"[DATABASE_TOOL_CREATE] Attempting to delete tool ID: {tool_id}")
    try:
        stmt = select(ToolRegistry).where(ToolRegistry.id == tool_id)
        res = await db.execute(stmt)
        tool = res.scalar_one_or_none()
        if not tool:
            return JSONResponse(
                status_code=status.HTTP_404_NOT_FOUND,
                content={"success": False, "message": "Tool not found"}
            )
            
        # Delete file from Supabase if downloadable
        if tool.type == "downloadable" and tool.file_path:
            from app.utils.supabase_storage import delete_file_from_supabase
            storage_path = tool.file_path
            if "storage/v1/object/public/" in storage_path:
                parts = storage_path.split(f"/{BUCKET_NAME}/")
                if len(parts) > 1:
                    storage_path = parts[1]
            await delete_file_from_supabase(storage_path)

        await db.delete(tool)
        await db.commit()
        print(f"[DATABASE_TOOL_CREATE] Tool ID: {tool_id} deleted successfully.")
        return JSONResponse(status_code=200, content={"success": True, "message": "Tool deleted"})
    except Exception as e:
        await db.rollback()
        print(f"[DATABASE_TOOL_CREATE] Exception deleting tool ID: {tool_id}: {str(e)}")
        return JSONResponse(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            content={"success": False, "message": f"Database error: {str(e)}"}
        )

@router.post("/tools/upload")
async def upload_tool_file(
    file: UploadFile = File(...),
    admin: User = Depends(get_current_admin)
):
    # Log [TOOL_UPLOAD] and [SUPABASE_UPLOAD]
    print(f"[TOOL_UPLOAD] Request received to upload file: '{file.filename}'")
    print(f"[SUPABASE_UPLOAD] Upload starting for file: '{file.filename}'")

    try:
        contents = await file.read()
        unique_id = uuid.uuid4()
        storage_filename = f"{unique_id}_{file.filename}"
        file_path = f"tools/{storage_filename}"
        
        # MIME mapping for spreadsheet uploads
        content_type = file.content_type or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        
        success, public_url, error_msg = await upload_file_to_supabase(
            file_bytes=contents,
            file_path=file_path,
            content_type=content_type
        )
        
        if not success:
            print(f"[SUPABASE_UPLOAD] Supabase storage upload rejected: {error_msg}")
            return JSONResponse(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                content={"success": False, "message": f"Supabase upload failed: {error_msg}"}
            )
            
        # Log [SUPABASE_URL_GENERATION]
        print(f"[SUPABASE_URL_GENERATION] Generated URL: '{public_url}' for file: '{file.filename}'")
        return {
            "success": True,
            "file_url": public_url, 
            "storage_path": file_path,
            "original_filename": file.filename,
            "storage_filename": storage_filename
        }
    except Exception as e:
        print(f"[TOOL_UPLOAD] Exception occurred during upload process: {str(e)}")
        return JSONResponse(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            content={"success": False, "message": f"Upload process error: {str(e)}"}
        )

@router.get("/tools/download/{tool_id}")
async def download_tool_file(
    tool_id: int,
    token: Optional[str] = None,
    db: AsyncSession = Depends(get_db)
):
    # Log [TOOL_DOWNLOAD]
    print(f"[TOOL_DOWNLOAD] Download proxy triggered for tool ID: {tool_id}")

    if token:
        payload = decode_access_token(token)
        if not payload:
            raise HTTPException(status_code=401, detail="Session expired or invalid token.")
        user_id = payload.get("user_id")
        if user_id:
            stmt_user = select(User).where(User.id == int(user_id))
            res_user = await db.execute(stmt_user)
            user = res_user.scalar_one_or_none()
            if user:
                if not user.is_active:
                    raise HTTPException(status_code=403, detail="User account is deactivated.")
                if user.role == "employee":
                    from app.models.employee_permission import EmployeePermission
                    perm_stmt = select(EmployeePermission).where(EmployeePermission.user_id == user.id)
                    perm_res = await db.execute(perm_stmt)
                    perm = perm_res.scalar_one_or_none()
                    if not perm or "wow_toolkit" not in (perm.allowed_tools or []):
                        raise HTTPException(status_code=403, detail="Access Denied. You do not have permission to access the WOW Toolkit.")

    try:
        stmt = select(ToolRegistry).where(ToolRegistry.id == tool_id)
        res = await db.execute(stmt)
        tool = res.scalar_one_or_none()
        
        if not tool or tool.type != "downloadable":
            raise HTTPException(status_code=404, detail="Tool not found or is not downloadable.")
            
        if not tool.file_path:
            raise HTTPException(status_code=404, detail="Requested tool does not contain a file path.")

        # Extract relative storage path
        storage_path = tool.file_path
        if "storage/v1/object/public/" in storage_path:
            parts = storage_path.split(f"/{BUCKET_NAME}/")
            if len(parts) > 1:
                storage_path = parts[1]

        # Log [SUPABASE_URL_GENERATION]
        import httpx
        headers = {
            "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}"
        }
        supabase_download_url = f"{SUPABASE_URL}/storage/v1/object/{BUCKET_NAME}/{storage_path.lstrip('/')}"
        print(f"[SUPABASE_URL_GENERATION] Proxied download URL: {supabase_download_url}")

        async def file_streamer():
            async with httpx.AsyncClient() as client:
                async with client.stream("GET", supabase_download_url, headers=headers, timeout=120.0) as stream_res:
                    if stream_res.status_code != 200:
                        print(f"[TOOL_DOWNLOAD] Supabase download error status code: {stream_res.status_code}")
                        raise HTTPException(status_code=400, detail="Failed to retrieve binary file from Supabase Storage.")
                    async for chunk in stream_res.aiter_bytes(chunk_size=16384):
                        yield chunk

        # Determine download filename
        download_filename = tool.original_filename or tool.name or "download"
        # Extract extension from tool file_path
        ext = ""
        if "." in tool.file_path:
            ext = "." + tool.file_path.split(".")[-1].lower()
            
        if ext and not download_filename.lower().endswith(ext):
            download_filename = download_filename + ext

        safe_filename = download_filename.replace('"', '\\"')
        
        # Detect media type
        ext_type = download_filename.split(".")[-1].lower() if "." in download_filename else "xlsx"
        media_types = {
            "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "xls": "application/vnd.ms-excel",
            "csv": "text/csv"
        }
        media_type = media_types.get(ext_type, "application/octet-stream")

        from fastapi.responses import StreamingResponse
        print(f"[TOOL_DOWNLOAD] Streaming file content for: '{download_filename}' with Content-Disposition")
        return StreamingResponse(
            file_streamer(),
            media_type=media_type,
            headers={
                "Content-Disposition": f'attachment; filename="{safe_filename}"',
                "X-Content-Type-Options": "nosniff"
            }
        )
    except Exception as e:
        print(f"[TOOL_DOWNLOAD] Exception in proxy download for tool ID {tool_id}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Proxy download failure: {str(e)}")

@router.get("/tools/preview/{tool_id}")
async def preview_tool_file(
    tool_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    # Log [TOOL_PREVIEW]
    print(f"[TOOL_PREVIEW] Preview requested for tool ID: {tool_id} by: {current_user.full_name}")

    try:
        stmt = select(ToolRegistry).where(ToolRegistry.id == tool_id)
        res = await db.execute(stmt)
        tool = res.scalar_one_or_none()
        
        if not tool or tool.type != "downloadable":
            raise HTTPException(status_code=404, detail="Tool not found or is not downloadable.")
            
        if not tool.file_path:
            raise HTTPException(status_code=404, detail="Tool has no associated spreadsheet file.")

        # Extract relative storage path
        storage_path = tool.file_path
        if "storage/v1/object/public/" in storage_path:
            parts = storage_path.split(f"/{BUCKET_NAME}/")
            if len(parts) > 1:
                storage_path = parts[1]

        # Fetch bytes from Supabase
        import httpx
        headers = {
            "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}"
        }
        supabase_download_url = f"{SUPABASE_URL}/storage/v1/object/{BUCKET_NAME}/{storage_path.lstrip('/')}"
        
        file_bytes = b""
        async with httpx.AsyncClient() as client:
            res_stream = await client.get(supabase_download_url, headers=headers, timeout=30.0)
            if res_stream.status_code == 200:
                file_bytes = res_stream.content
            else:
                print(f"[TOOL_PREVIEW] Failed to download file for parsing: {res_stream.status_code}")
                raise HTTPException(status_code=400, detail="Failed to retrieve file from storage for preview.")

        # Parse file contents dynamically
        import io
        sheets_data = []
        file_ext = storage_path.split(".")[-1].lower() if "." in storage_path else ""
        
        if file_ext == "xlsx":
            try:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
                for sheet_name in wb.sheetnames[:3]: # Limit to first 3 sheets to keep response light
                    sheet = wb[sheet_name]
                    grid = []
                    # Get first 15 rows and first 10 columns
                    max_r = min(sheet.max_row, 15)
                    max_c = min(sheet.max_column, 10)
                    
                    if max_r == 0 or max_c == 0:
                        continue
                        
                    for r in range(1, max_r + 1):
                        row_data = []
                        for c in range(1, max_c + 1):
                            val = sheet.cell(row=r, column=c).value
                            row_data.append(str(val) if val is not None else "")
                        grid.append(row_data)
                    sheets_data.append({
                        "name": sheet_name,
                        "data": grid
                    })
            except Exception as e:
                print(f"[TOOL_PREVIEW] Excel openpyxl parsing exception: {str(e)}")
                
        elif file_ext == "csv":
            try:
                import csv
                text_content = file_bytes.decode("utf-8", errors="ignore")
                reader = csv.reader(io.StringIO(text_content))
                grid = []
                for idx, row in enumerate(reader):
                    if idx >= 15:
                        break
                    grid.append(row[:10])
                sheets_data.append({
                    "name": "CSV Data",
                    "data": grid
                })
            except Exception as e:
                print(f"[TOOL_PREVIEW] CSV parsing exception: {str(e)}")

        # Fallback if no sheets parsed or parsing failed
        if not sheets_data:
            sheets_data.append({
                "name": "Preview",
                "data": [
                    ["File Name", tool.original_filename or "Template file"],
                    ["File Type", file_ext.upper()],
                    ["File Size", f"{len(file_bytes) / 1024:.1f} KB"],
                    ["Upload Date", tool.created_at.strftime("%Y-%m-%d") if tool.created_at else "N/A"]
                ]
            })

        return {
            "success": True,
            "name": tool.name,
            "original_filename": tool.original_filename,
            "upload_date": tool.created_at.isoformat() if tool.created_at else None,
            "file_type": file_ext.upper(),
            "sheets": sheets_data
        }
    except Exception as e:
        print(f"[TOOL_PREVIEW] Exception parsing preview for tool ID {tool_id}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Preview parser failed: {str(e)}")


# --- Calculator 1: Retirement Age Predictor ---

class RetirementInput(BaseModel):
    current_age: float
    expected_retirement_age: float
    life_expectancy: float
    current_monthly_expenses: float
    expected_inflation_rate: float
    current_monthly_income: float
    savings_rate: float
    expected_investment_return: float
    post_retirement_return: float

class SensitivityPoint(BaseModel):
    age: float
    corpus_needed: float

class RetirementResult(BaseModel):
    years_to_retirement: float
    years_in_retirement: float
    monthly_savings: float
    monthly_expenses_at_retirement: float
    required_retirement_corpus: float
    savings_corpus_at_retirement: float
    corpus_surplus_deficit: float
    track_status: str
    sensitivity_table: List[SensitivityPoint]

@router.post("/retirement/calculate", response_model=RetirementResult)
def calculate_retirement(
    inputs: RetirementInput,
    current_user: User = Depends(get_current_active_user)
):
    if inputs.current_age >= inputs.expected_retirement_age:
        raise HTTPException(status_code=400, detail="Current age must be less than expected retirement age.")
    if inputs.expected_retirement_age >= inputs.life_expectancy:
        raise HTTPException(status_code=400, detail="Expected retirement age must be less than life expectancy.")
    
    results = calculate_retirement_predictor(
        current_age=inputs.current_age,
        expected_retirement_age=inputs.expected_retirement_age,
        life_expectancy=inputs.life_expectancy,
        current_monthly_expenses=inputs.current_monthly_expenses,
        expected_inflation_rate=inputs.expected_inflation_rate,
        current_monthly_income=inputs.current_monthly_income,
        savings_rate=inputs.savings_rate,
        expected_investment_return=inputs.expected_investment_return,
        post_retirement_return=inputs.post_retirement_return
    )
    
    return RetirementResult(
        years_to_retirement=results["years_to_retirement"],
        years_in_retirement=results["years_in_retirement"],
        monthly_savings=results["monthly_savings"],
        monthly_expenses_at_retirement=results["monthly_expenses_at_retirement"],
        required_retirement_corpus=results["required_retirement_corpus"],
        savings_corpus_at_retirement=results["savings_corpus_at_retirement"],
        corpus_surplus_deficit=results["corpus_surplus_deficit"],
        track_status=results["track_status"],
        sensitivity_table=[SensitivityPoint(age=p["age"], corpus_needed=p["corpus_needed"]) for p in results["sensitivity_table"]]
    )


# --- Calculator 2: Cost of Delay ---

class DelayInput(BaseModel):
    monthly_sip_amount: float
    expected_annual_return: float
    target_age: float
    current_age: float

class DelayPoint(BaseModel):
    start_age: float
    years_to_invest: float
    total_invested: float
    corpus_at_target: float
    vs_starting_at_25: float
    delay_cost_percent: float

class DelayResult(BaseModel):
    warning_text: str
    delay_table: List[DelayPoint]

@router.post("/cost-delay/calculate", response_model=DelayResult)
def calculate_delay(
    inputs: DelayInput,
    current_user: User = Depends(get_current_active_user)
):
    if inputs.monthly_sip_amount <= 0:
        raise HTTPException(status_code=400, detail="Monthly SIP Amount must be greater than zero.")
    if inputs.expected_annual_return < 0 or inputs.expected_annual_return > 1:
        raise HTTPException(status_code=400, detail="Expected Annual Return must be between 0% and 100% (0.0 and 1.0).")
    if inputs.current_age < 0 or inputs.target_age < 0:
        raise HTTPException(status_code=400, detail="Age values cannot be negative.")
    if inputs.current_age >= inputs.target_age:
        raise HTTPException(status_code=400, detail="Current age must be less than the target age.")
        
    results = calculate_cost_of_delay(
        monthly_sip_amount=inputs.monthly_sip_amount,
        expected_annual_return=inputs.expected_annual_return,
        target_age=inputs.target_age,
        current_age=inputs.current_age
    )
    return DelayResult(
        warning_text=results["warning_text"],
        delay_table=[DelayPoint(**p) for p in results["delay_table"]]
    )


# --- Calculator 3: SIP + Home Loan Impact ---

class SipLoanInput(BaseModel):
    monthly_sip: float
    sip_return: float
    sip_duration: float
    stepup_rate: float
    loan_amount: float
    loan_rate: float
    loan_tenure: float
    down_payment: float
    appreciation_rate: float
    tax_benefit: float

class SipSeriesPoint(BaseModel):
    year: int
    simple_balance: float
    simple_invested: float
    stepup_balance: float
    stepup_invested: float

class LoanSeriesPoint(BaseModel):
    year: int
    remaining_balance: float
    principal_paid: float
    interest_paid: float
    property_value: float

class SipLoanResult(BaseModel):
    simple_sip_corpus: float
    total_amount_invested: float
    wealth_gain: float
    stepup_sip_corpus: float
    return_multiple: float
    monthly_emi: float
    total_amount_paid: float
    total_interest_paid: float
    property_value_maturity: float
    net_property_gain: float
    sip_emi_invested: float
    sip_advantage_over_property: float
    total_monthly_outflow: float
    combined_sip_corpus: float
    combined_property_net: float
    combined_wealth: float
    effective_emi: float
    recommendation_msg: str
    sip_series: List[SipSeriesPoint]
    loan_series: List[LoanSeriesPoint]

@router.post("/sip-home-loan/calculate", response_model=SipLoanResult)
def calculate_sip_loan(
    inputs: SipLoanInput,
    current_user: User = Depends(get_current_active_user)
):
    if inputs.monthly_sip < 0 or inputs.loan_amount < 0 or inputs.down_payment < 0:
        raise HTTPException(status_code=400, detail="Financial amounts cannot be negative.")
    if inputs.sip_duration <= 0 or inputs.loan_tenure <= 0:
        raise HTTPException(status_code=400, detail="Tenures/durations must be greater than zero.")
    if inputs.sip_return < 0 or inputs.loan_rate < 0 or inputs.stepup_rate < 0 or inputs.appreciation_rate < 0:
        raise HTTPException(status_code=400, detail="Rates/returns cannot be negative.")
        
    results = calculate_sip_home_loan_impact(
        monthly_sip=inputs.monthly_sip,
        sip_return=inputs.sip_return,
        sip_duration=inputs.sip_duration,
        stepup_rate=inputs.stepup_rate,
        loan_amount=inputs.loan_amount,
        loan_rate=inputs.loan_rate,
        loan_tenure=inputs.loan_tenure,
        down_payment=inputs.down_payment,
        appreciation_rate=inputs.appreciation_rate,
        tax_benefit=inputs.tax_benefit
    )
    return SipLoanResult(**results)


# --- Calculator 4: Financial Freedom Date ---

class FreedomDateInput(BaseModel):
    current_age: float
    birth_year: float
    current_monthly_expenses: float
    expected_inflation: float
    annual_investment_return: float
    withdrawal_rate: float
    current_net_worth: float
    monthly_savings: float
    stepup_rate: float

class TimelinePoint(BaseModel):
    year: int
    age: int
    simple_net_worth: float
    stepup_net_worth: float
    fi_target: float

class FreedomDateResult(BaseModel):
    fi_number: float
    fi_number_inflation_20: float
    years_to_fi: float
    years_to_fi_stepup: float
    fi_achievement_year: float
    fi_age_at_achievement: float
    fi_target: float
    current_net_worth_val: float
    remaining_gap: float
    percent_fi_achieved: float
    years_remaining: float
    monthly_saved: float
    progress_milestone: str
    annual_withdrawal: float
    monthly_income_at_fi: float
    monthly_income_inflation_10: float
    monthly_income_inflation_20: float
    safe_fi_buffer: float
    freedom_date_message: str
    timeline_series: List[TimelinePoint]

@router.post("/freedom-date/calculate", response_model=FreedomDateResult)
def calculate_freedom(
    inputs: FreedomDateInput,
    current_user: User = Depends(get_current_active_user)
):
    results = calculate_freedom_date(
        current_age=inputs.current_age,
        birth_year=inputs.birth_year,
        current_monthly_expenses=inputs.current_monthly_expenses,
        expected_inflation=inputs.expected_inflation,
        annual_investment_return=inputs.annual_investment_return,
        withdrawal_rate=inputs.withdrawal_rate,
        current_net_worth=inputs.current_net_worth,
        monthly_savings=inputs.monthly_savings,
        stepup_rate=inputs.stepup_rate
    )
    return FreedomDateResult(**results)


# --- Calculator 5: Goal Visualization Dashboard ---

class GoalItem(BaseModel):
    name: str
    target_amount: float
    current_saved: float
    monthly_sip: float
    timeline_years: float

class CalculatedGoalItem(BaseModel):
    id: Optional[int] = None
    name: str
    target_amount: float
    current_saved: float
    monthly_sip: float
    timeline_years: float
    percent_achieved: float
    status: str

class GoalDashboardResult(BaseModel):
    goals: List[CalculatedGoalItem]
    total_target: float
    total_saved: float
    total_sip: float
    overall_percent_achieved: float

@router.post("/goal-dashboard/calculate", response_model=GoalDashboardResult)
def calculate_goals(inputs: List[GoalItem]):
    goals_list = [g.model_dump() for g in inputs]
    results = calculate_goal_dashboard(goals_list)
    return GoalDashboardResult(
        goals=[CalculatedGoalItem(**g) for g in results["goals"]],
        total_target=results["total_target"],
        total_saved=results["total_saved"],
        total_sip=results["total_sip"],
        overall_percent_achieved=results["overall_percent_achieved"]
    )


# --- PERSISTED DATABASE APIS: Financial Goals & Family Vault ---

from app.models.wow import (
    FinancialGoal,
    VaultFamilyMember,
    VaultNominee,
    VaultInsurancePolicy,
    VaultLoan,
    VaultInvestment,
    VaultImportantDocument,
    VaultEmergencyContact,
    VaultBankAccount,
    WOWUserInputs
)

# Schemas for Vault Sections
class VaultFamilyMemberSchema(BaseModel):
    id: Optional[int] = None
    name: str
    relationship: str
    dob: str
    pan_number: Optional[str] = None
    aadhaar_last_four: Optional[str] = None
    blood_group: Optional[str] = None

class VaultNomineeSchema(BaseModel):
    id: Optional[int] = None
    name: str
    relationship: str
    allocation_percentage: float
    associated_asset: Optional[str] = None

class VaultInsurancePolicySchema(BaseModel):
    id: Optional[int] = None
    policy_type: str
    company: str
    policy_number: str
    sum_assured: float
    premium_amount: float
    expiry_date: Optional[str] = None

class VaultLoanSchema(BaseModel):
    id: Optional[int] = None
    loan_name: str
    lender: str
    outstanding_amount: float
    emi: float
    interest_rate: float

class VaultInvestmentSchema(BaseModel):
    id: Optional[int] = None
    investment_type: str
    scheme_name: str
    account_folio_number: str
    current_value: float
    nominee: str
    institution: str

class VaultImportantDocumentSchema(BaseModel):
    id: Optional[int] = None
    document_name: str
    storage_location: str
    last_updated: str
    digital_copy_stored_at: str
    status: str

class VaultEmergencyContactSchema(BaseModel):
    id: Optional[int] = None
    name: str
    relationship: str
    mobile: str
    email: str
    role_purpose: str

class VaultBankAccountSchema(BaseModel):
    id: Optional[int] = None
    bank_card_name: str
    account_type: str
    last_four_digits: str
    branch_limit: str
    nominee: str
    status: str

class VaultItemPayload(BaseModel):
    family_member: Optional[VaultFamilyMemberSchema] = None
    nominee: Optional[VaultNomineeSchema] = None
    insurance_policy: Optional[VaultInsurancePolicySchema] = None
    loan: Optional[VaultLoanSchema] = None
    investment: Optional[VaultInvestmentSchema] = None
    important_document: Optional[VaultImportantDocumentSchema] = None
    emergency_contact: Optional[VaultEmergencyContactSchema] = None
    bank_account: Optional[VaultBankAccountSchema] = None

def get_vault_model(type_str: str):
    mapping = {
        "family": VaultFamilyMember,
        "nominee": VaultNominee,
        "insurance": VaultInsurancePolicy,
        "loan": VaultLoan,
        "investment": VaultInvestment,
        "document": VaultImportantDocument,
        "contact": VaultEmergencyContact,
        "bank_account": VaultBankAccount
    }
    model = mapping.get(type_str.lower())
    if not model:
        raise HTTPException(status_code=400, detail=f"Invalid vault item type: {type_str}")
    return model


# --- Goals CRUD ---

class CreateGoalPayload(BaseModel):
    name: str
    target_amount: float
    current_saved: float
    monthly_sip: float
    timeline_years: float

@router.get("/goals")
async def get_goals(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    stmt = select(FinancialGoal).where(FinancialGoal.user_id == current_user.id)
    res = await db.execute(stmt)
    db_goals = res.scalars().all()
    goals_list = [
        {
            "id": g.id,
            "name": g.name,
            "target_amount": g.target_amount,
            "current_saved": g.current_saved,
            "monthly_sip": g.monthly_sip,
            "timeline_years": g.timeline_years
        }
        for g in db_goals
    ]
    calc_results = calculate_goal_dashboard(goals_list)
    for i, g in enumerate(calc_results["goals"]):
        g["id"] = db_goals[i].id
    return calc_results

@router.post("/goals")
async def create_goal(
    payload: CreateGoalPayload,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    goal = FinancialGoal(
        user_id=current_user.id,
        name=payload.name,
        target_amount=payload.target_amount,
        current_saved=payload.current_saved,
        monthly_sip=payload.monthly_sip,
        timeline_years=payload.timeline_years
    )
    db.add(goal)
    await db.commit()
    await db.refresh(goal)
    return goal

@router.put("/goals/{id}")
async def update_goal(
    id: int,
    payload: CreateGoalPayload,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    stmt = select(FinancialGoal).where(FinancialGoal.id == id, FinancialGoal.user_id == current_user.id)
    res = await db.execute(stmt)
    goal = res.scalar_one_or_none()
    if not goal:
        raise HTTPException(status_code=404, detail="Goal not found")
    goal.name = payload.name
    goal.target_amount = payload.target_amount
    goal.current_saved = payload.current_saved
    goal.monthly_sip = payload.monthly_sip
    goal.timeline_years = payload.timeline_years
    await db.commit()
    await db.refresh(goal)
    return goal

@router.delete("/goals/{id}")
async def delete_goal(
    id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    stmt = select(FinancialGoal).where(FinancialGoal.id == id, FinancialGoal.user_id == current_user.id)
    res = await db.execute(stmt)
    goal = res.scalar_one_or_none()
    if not goal:
        raise HTTPException(status_code=404, detail="Goal not found")
    await db.delete(goal)
    await db.commit()
    return {"message": "Goal deleted successfully"}


# --- Vault CRUD ---

@router.get("/vault")
async def get_vault(
    user_id: Optional[int] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    try:
        target_user_id = current_user.id
        if user_id and current_user.role == "admin":
            target_user_id = user_id

        res_family = await db.execute(select(VaultFamilyMember).where(VaultFamilyMember.user_id == target_user_id))
        res_nominee = await db.execute(select(VaultNominee).where(VaultNominee.user_id == target_user_id))
        res_insurance = await db.execute(select(VaultInsurancePolicy).where(VaultInsurancePolicy.user_id == target_user_id))
        res_loan = await db.execute(select(VaultLoan).where(VaultLoan.user_id == target_user_id))
        res_investment = await db.execute(select(VaultInvestment).where(VaultInvestment.user_id == target_user_id))
        res_document = await db.execute(select(VaultImportantDocument).where(VaultImportantDocument.user_id == target_user_id))
        res_contact = await db.execute(select(VaultEmergencyContact).where(VaultEmergencyContact.user_id == target_user_id))
        res_bank = await db.execute(select(VaultBankAccount).where(VaultBankAccount.user_id == target_user_id))

        return {
            "family_members": res_family.scalars().all(),
            "nominees": res_nominee.scalars().all(),
            "insurance_policies": res_insurance.scalars().all(),
            "loans": res_loan.scalars().all(),
            "investments": res_investment.scalars().all(),
            "important_documents": res_document.scalars().all(),
            "emergency_contacts": res_contact.scalars().all(),
            "bank_accounts": res_bank.scalars().all()
        }
    except Exception as e:
        import traceback
        print(f"[WOW VAULT ERROR] Error in get_vault: {str(e)}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Internal database error in vault: {str(e)}")

@router.post("/vault")
async def create_vault_item(
    type: str,
    payload: VaultItemPayload,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    model_class = get_vault_model(type)
    
    item_data = None
    if type == "family" and payload.family_member:
        item_data = payload.family_member.model_dump(exclude_unset=True)
    elif type == "nominee" and payload.nominee:
        item_data = payload.nominee.model_dump(exclude_unset=True)
    elif type == "insurance" and payload.insurance_policy:
        item_data = payload.insurance_policy.model_dump(exclude_unset=True)
    elif type == "loan" and payload.loan:
        item_data = payload.loan.model_dump(exclude_unset=True)
    elif type == "investment" and payload.investment:
        item_data = payload.investment.model_dump(exclude_unset=True)
    elif type == "document" and payload.important_document:
        item_data = payload.important_document.model_dump(exclude_unset=True)
    elif type == "contact" and payload.emergency_contact:
        item_data = payload.emergency_contact.model_dump(exclude_unset=True)
    elif type == "bank_account" and payload.bank_account:
        item_data = payload.bank_account.model_dump(exclude_unset=True)
        
    if not item_data:
        raise HTTPException(status_code=400, detail=f"Payload for type {type} is missing.")
        
    item_data.pop("id", None)
    
    db_item = model_class(user_id=current_user.id, **item_data)
    db.add(db_item)
    await db.commit()
    await db.refresh(db_item)
    return db_item

@router.put("/vault/{id}")
async def update_vault_item(
    id: int,
    type: str,
    payload: VaultItemPayload,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    model_class = get_vault_model(type)
    stmt = select(model_class).where(model_class.id == id)
    if current_user.role != "admin":
        stmt = stmt.where(model_class.user_id == current_user.id)
        
    res = await db.execute(stmt)
    db_item = res.scalar_one_or_none()
    if not db_item:
        raise HTTPException(status_code=404, detail="Vault item not found")
        
    item_data = None
    if type == "family" and payload.family_member:
        item_data = payload.family_member.model_dump(exclude_unset=True)
    elif type == "nominee" and payload.nominee:
        item_data = payload.nominee.model_dump(exclude_unset=True)
    elif type == "insurance" and payload.insurance_policy:
        item_data = payload.insurance_policy.model_dump(exclude_unset=True)
    elif type == "loan" and payload.loan:
        item_data = payload.loan.model_dump(exclude_unset=True)
    elif type == "investment" and payload.investment:
        item_data = payload.investment.model_dump(exclude_unset=True)
    elif type == "document" and payload.important_document:
        item_data = payload.important_document.model_dump(exclude_unset=True)
    elif type == "contact" and payload.emergency_contact:
        item_data = payload.emergency_contact.model_dump(exclude_unset=True)
    elif type == "bank_account" and payload.bank_account:
        item_data = payload.bank_account.model_dump(exclude_unset=True)
        
    if not item_data:
        raise HTTPException(status_code=400, detail=f"Payload for type {type} is missing.")
        
    for k, v in item_data.items():
        if k != "id" and k != "user_id":
            setattr(db_item, k, v)
            
    await db.commit()
    await db.refresh(db_item)
    return db_item

@router.delete("/vault/{id}")
async def delete_vault_item(
    id: int,
    type: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    model_class = get_vault_model(type)
    stmt = select(model_class).where(model_class.id == id)
    if current_user.role != "admin":
        stmt = stmt.where(model_class.user_id == current_user.id)
        
    res = await db.execute(stmt)
    db_item = res.scalar_one_or_none()
    if not db_item:
        raise HTTPException(status_code=404, detail="Vault item not found")
        
    await db.delete(db_item)
    await db.commit()
    return {"message": "Vault item deleted successfully"}


# --- WOW User Inputs CRUD (Autosave / Data Persistence) ---

class WOWUserInputsSchema(BaseModel):
    retirement_inputs: Optional[dict] = None
    cost_of_delay_inputs: Optional[dict] = None
    sip_home_loan_inputs: Optional[dict] = None
    freedom_date_inputs: Optional[dict] = None

@router.get("/inputs", response_model=WOWUserInputsSchema)
async def get_wow_inputs(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    stmt = select(WOWUserInputs).where(WOWUserInputs.user_id == current_user.id)
    res = await db.execute(stmt)
    db_inputs = res.scalar_one_or_none()
    if not db_inputs:
        return {
            "retirement_inputs": {},
            "cost_of_delay_inputs": {},
            "sip_home_loan_inputs": {},
            "freedom_date_inputs": {}
        }
    return {
        "retirement_inputs": db_inputs.retirement_inputs or {},
        "cost_of_delay_inputs": db_inputs.cost_of_delay_inputs or {},
        "sip_home_loan_inputs": db_inputs.sip_home_loan_inputs or {},
        "freedom_date_inputs": db_inputs.freedom_date_inputs or {}
    }

@router.post("/inputs")
async def save_wow_inputs(
    payload: WOWUserInputsSchema,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    stmt = select(WOWUserInputs).where(WOWUserInputs.user_id == current_user.id)
    res = await db.execute(stmt)
    db_inputs = res.scalar_one_or_none()
    if not db_inputs:
        db_inputs = WOWUserInputs(user_id=current_user.id)
        db.add(db_inputs)
        
    if payload.retirement_inputs is not None:
        db_inputs.retirement_inputs = payload.retirement_inputs
    if payload.cost_of_delay_inputs is not None:
        db_inputs.cost_of_delay_inputs = payload.cost_of_delay_inputs
    if payload.sip_home_loan_inputs is not None:
        db_inputs.sip_home_loan_inputs = payload.sip_home_loan_inputs
    if payload.freedom_date_inputs is not None:
        db_inputs.freedom_date_inputs = payload.freedom_date_inputs
        
    await db.commit()
    await db.refresh(db_inputs)
    return {"status": "success"}

