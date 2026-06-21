import openpyxl
import pandas as pd

wb = openpyxl.load_workbook('./Wow_Tool_Fixed (1) (1).xlsx', data_only=False)
wb_val = openpyxl.load_workbook('./Wow_Tool_Fixed (1) (1).xlsx', data_only=True)

for sheet_name in wb.sheetnames:
    print(f"\n================= SHEET: {sheet_name} =================")
    sheet = wb[sheet_name]
    sheet_val = wb_val[sheet_name]
    
    # Let's collect cells that have values or formulas
    rows = []
    for r in range(1, 100):
        for c in range(1, 26):
            cell = sheet.cell(row=r, column=c)
            cell_val = sheet_val.cell(row=r, column=c)
            if cell.value is not None:
                val = cell_val.value
                formula = cell.value if str(cell.value).startswith('=') else None
                col_letter = openpyxl.utils.get_column_letter(c)
                rows.append({
                    "Cell": f"{col_letter}{r}",
                    "Label/Value": val if formula is None else f"Formula output: {val}",
                    "Formula": formula,
                    "RawValue": cell.value
                })
    df = pd.DataFrame(rows)
    # Print first 120 cells
    pd.set_option('display.max_rows', 200)
    pd.set_option('display.max_colwidth', None)
    print(df.to_string())
