import openpyxl
import os

wb = openpyxl.load_workbook('/home/akhila/Downloads/Learning_Platform_App/Wow_Tool_Fixed (1) (1).xlsx', data_only=False)
wb_val = openpyxl.load_workbook('/home/akhila/Downloads/Learning_Platform_App/Wow_Tool_Fixed (1) (1).xlsx', data_only=True)

for sheetname in wb.sheetnames:
    print(f"=== Sheet: {sheetname} ===")
    sheet_formula = wb[sheetname]
    sheet_value = wb_val[sheetname]
    for row in range(1, 100):
        for col in range(1, 15):
            cell_formula = sheet_formula.cell(row=row, column=col)
            cell_value = sheet_value.cell(row=row, column=col)
            if cell_formula.value is not None or cell_value.value is not None:
                coord = cell_formula.coordinate
                print(f"{coord} | Formula: {cell_formula.value} | Value: {cell_value.value}")
    print("\n")
